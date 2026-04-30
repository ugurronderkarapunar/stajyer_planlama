"""
=========================================================
  Stajyer Takip ve Puantaj Yönetim Sistemi
  Versiyon: 1.0
=========================================================
"""

import streamlit as st
import pandas as pd
from sqlalchemy import create_engine, text, MetaData, Table, Column, Integer, String, ForeignKey
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
import holidays as holidays_lib
from datetime import date, timedelta
import calendar
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════
#  SAYFA YAPILANDIRMASI
# ══════════════════════════════════════════════════════
st.set_page_config(
    page_title="Stajyer Takip Sistemi",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════
#  SABİTLER
# ══════════════════════════════════════════════════════
PERIYOT_OPTIONS = [
    "Hafta İçi Her Gün",
    "Pazartesi-Salı-Çarşamba",
    "Çarşamba-Perşembe-Cuma",
]

BOLUM_OPTIONS = ["Makineci", "Güverte"]
IZIN_TURU_OPTIONS = ["Raporlu", "Raporsuz"]

# Hafta içi gün numaraları (0=Pazartesi … 6=Pazar)
PERIYOT_DAYS: dict[str, list[int]] = {
    "Hafta İçi Her Gün": [0, 1, 2, 3, 4],
    "Pazartesi-Salı-Çarşamba": [0, 1, 2],
    "Çarşamba-Perşembe-Cuma": [2, 3, 4],
}

AY_ADLARI = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

GUN_ADLARI = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]

# ══════════════════════════════════════════════════════
#  VERİTABANI — SİNGLETON BAĞLANTI
# ══════════════════════════════════════════════════════

@st.cache_resource(show_spinner="Veritabanına bağlanılıyor…")
def get_engine():
    """
    Önce st.secrets içindeki DATABASE_URL'yi dener (Supabase / PostgreSQL).
    Bulunamazsa yerel SQLite'a düşer.
    """
    try:
        db_url = st.secrets["DATABASE_URL"]
    except (KeyError, FileNotFoundError):
        db_url = "sqlite:///./stajyer_takip.db"

    connect_args = {"check_same_thread": False} if "sqlite" in str(db_url) else {}
    engine = create_engine(db_url, connect_args=connect_args, pool_pre_ping=True)
    return engine


@st.cache_resource
def get_metadata():
    """SQLAlchemy tablo meta-verisini döndürür (diyalekt-bağımsız)."""
    meta = MetaData()

    Table(
        "stajyerler", meta,
        Column("id",                Integer, primary_key=True, autoincrement=True),
        Column("ad",                String(100), nullable=False),
        Column("soyad",             String(100), nullable=False),
        Column("okul",              String(200)),
        Column("telefon",           String(30)),
        Column("sicil_no",          String(50),  unique=True),
        Column("staj_gemisi",       String(100)),
        Column("bolum",             String(50)),
        Column("calisma_periyodu",  String(100)),
        extend_existing=True,
    )

    Table(
        "izinler", meta,
        Column("id",                Integer, primary_key=True, autoincrement=True),
        Column("stajyer_id",        Integer, ForeignKey("stajyerler.id", ondelete="CASCADE")),
        Column("baslangic_tarihi",  String(20), nullable=False),
        Column("bitis_tarihi",      String(20), nullable=False),
        Column("izin_turu",         String(50)),
        extend_existing=True,
    )

    return meta


def init_db():
    """Tablolar yoksa oluşturur."""
    try:
        meta = get_metadata()
        meta.create_all(get_engine())
    except SQLAlchemyError as exc:
        st.error(f"Veritabanı başlatılamadı: {exc}")
        st.stop()


# ══════════════════════════════════════════════════════
#  VERİ ERİŞİM FONKSİYONLARI
# ══════════════════════════════════════════════════════

def get_all_stajyerler() -> pd.DataFrame:
    with get_engine().connect() as conn:
        result = conn.execute(
            text("SELECT * FROM stajyerler ORDER BY staj_gemisi, soyad, ad")
        )
        rows = result.fetchall()
        return pd.DataFrame(rows, columns=list(result.keys())) if rows else pd.DataFrame()


def add_stajyer(data: dict) -> None:
    with get_engine().connect() as conn:
        conn.execute(
            text("""
                INSERT INTO stajyerler
                    (ad, soyad, okul, telefon, sicil_no, staj_gemisi, bolum, calisma_periyodu)
                VALUES
                    (:ad, :soyad, :okul, :telefon, :sicil_no, :staj_gemisi, :bolum, :calisma_periyodu)
            """),
            data,
        )
        conn.commit()


def delete_stajyer(stajyer_id: int) -> None:
    with get_engine().connect() as conn:
        conn.execute(text("DELETE FROM izinler   WHERE stajyer_id = :sid"), {"sid": stajyer_id})
        conn.execute(text("DELETE FROM stajyerler WHERE id = :sid"),         {"sid": stajyer_id})
        conn.commit()


def get_izinler_for_month(stajyer_id: int, year: int, month: int) -> pd.DataFrame:
    num_days = calendar.monthrange(year, month)[1]
    start    = f"{year}-{month:02d}-01"
    end      = f"{year}-{month:02d}-{num_days:02d}"
    with get_engine().connect() as conn:
        result = conn.execute(
            text("""
                SELECT * FROM izinler
                WHERE stajyer_id       = :sid
                  AND bitis_tarihi    >= :start
                  AND baslangic_tarihi<= :end
            """),
            {"sid": stajyer_id, "start": start, "end": end},
        )
        rows = result.fetchall()
        return pd.DataFrame(rows, columns=list(result.keys())) if rows else pd.DataFrame()


def get_all_izinler() -> pd.DataFrame:
    with get_engine().connect() as conn:
        result = conn.execute(
            text("""
                SELECT i.id, s.ad, s.soyad, s.staj_gemisi,
                       i.baslangic_tarihi, i.bitis_tarihi, i.izin_turu
                FROM izinler   i
                JOIN stajyerler s ON i.stajyer_id = s.id
                ORDER BY i.baslangic_tarihi DESC
            """)
        )
        rows = result.fetchall()
        return pd.DataFrame(rows, columns=list(result.keys())) if rows else pd.DataFrame()


def add_izin(data: dict) -> None:
    with get_engine().connect() as conn:
        conn.execute(
            text("""
                INSERT INTO izinler (stajyer_id, baslangic_tarihi, bitis_tarihi, izin_turu)
                VALUES (:stajyer_id, :baslangic_tarihi, :bitis_tarihi, :izin_turu)
            """),
            data,
        )
        conn.commit()


def delete_izin(izin_id: int) -> None:
    with get_engine().connect() as conn:
        conn.execute(text("DELETE FROM izinler WHERE id = :iid"), {"iid": izin_id})
        conn.commit()


# ══════════════════════════════════════════════════════
#  İŞ MANTIĞI
# ══════════════════════════════════════════════════════

def get_tr_holidays(year: int) -> dict:
    """Türkiye resmi tatillerini {date: isim} olarak döndürür."""
    return dict(holidays_lib.country_holidays("TR", years=year))


def get_day_status(
    day_date: date,
    periyot: str,
    izinler_df: pd.DataFrame,
    tr_holidays_dict: dict,
) -> str:
    """
    Dönen değerler:
      HAFTA SONU | TATİL | - | RAPORLU | RAPORSUZ DEVAMSIZLIK | 1
    """
    weekday   = day_date.weekday()
    work_days = PERIYOT_DAYS.get(periyot, [0, 1, 2, 3, 4])

    if weekday >= 5:
        return "HAFTA SONU"

    if day_date in tr_holidays_dict:
        return "TATİL"

    if weekday not in work_days:
        return "-"

    if not izinler_df.empty:
        for _, izin in izinler_df.iterrows():
            try:
                start = date.fromisoformat(str(izin["baslangic_tarihi"])[:10])
                end   = date.fromisoformat(str(izin["bitis_tarihi"])[:10])
            except ValueError:
                continue
            if start <= day_date <= end:
                return "RAPORLU" if str(izin["izin_turu"]) == "Raporlu" else "RAPORSUZ DEVAMSIZLIK"

    return "1"


def validate_leave_dates(
    start_date: date, end_date: date, periyot: str
) -> tuple[bool, list[str]]:
    """
    Dönüş:
      (izin aralığında en az 1 çalışma günü var mı, çalışma dışı gün listesi)
    """
    work_days = PERIYOT_DAYS.get(periyot, [0, 1, 2, 3, 4])
    has_work_day = False
    non_work_days: list[str] = []
    current = start_date
    while current <= end_date:
        if current.weekday() in work_days:
            has_work_day = True
        else:
            non_work_days.append(current.strftime("%d.%m.%Y"))
        current += timedelta(days=1)
    return has_work_day, non_work_days


# ══════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════

# Renk sabitleri
C_HEADER   = "1F4E79"
C_SUBHDR   = "2E75B6"
C_WEEKEND  = "D0CECE"
C_HOLIDAY  = "C00000"
C_WORKDAY  = "70AD47"
C_RAPORLU  = "FFD966"
C_RAPORSUZ = "F4B183"
C_NODAY    = "EDEDED"
C_TOTAL    = "BDD7EE"
C_WHITE    = "FFFFFF"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color="000000", size=9) -> Font:
    return Font(bold=bold, color=color, size=size)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def create_excel(year: int, month: int, stajyerler_df: pd.DataFrame) -> BytesIO:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"{AY_ADLARI[month - 1]} {year}"

    tr_hols  = get_tr_holidays(year)
    num_days = calendar.monthrange(year, month)[1]

    # ── Satır 1: Başlık ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    static_cols  = ["Ad Soyad", "Gemi", "Bölüm", "Periyot"]
    day_col_start = len(static_cols) + 1          # Excel sütun indisi (1-tabanlı)
    total_col     = day_col_start + num_days

    for ci, hdr in enumerate(static_cols, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill      = _fill(C_HEADER)
        c.font      = _font(bold=True, color=C_WHITE, size=10)
        c.alignment = _center()
        c.border    = _thin_border()

    for d in range(1, num_days + 1):
        ci = day_col_start + d - 1
        c  = ws.cell(row=1, column=ci, value=d)
        c.fill      = _fill(C_HEADER)
        c.font      = _font(bold=True, color=C_WHITE, size=9)
        c.alignment = _center()
        c.border    = _thin_border()

    tc = ws.cell(row=1, column=total_col, value="Toplam\nGün")
    tc.fill      = _fill(C_SUBHDR)
    tc.font      = _font(bold=True, color=C_WHITE, size=10)
    tc.alignment = _center()
    tc.border    = _thin_border()

    # ── Satır 2: Gün adları ──────────────────────────────────────────────────
    for d in range(1, num_days + 1):
        ci       = day_col_start + d - 1
        day_date = date(year, month, d)
        day_name = GUN_ADLARI[day_date.weekday()]
        c        = ws.cell(row=2, column=ci, value=day_name)
        c.alignment = _center()
        c.border    = _thin_border()
        if day_date.weekday() >= 5:
            c.fill = _fill(C_WEEKEND)
            c.font = _font(bold=True, color="595959", size=8)
        elif day_date in tr_hols:
            c.fill = _fill(C_HOLIDAY)
            c.font = _font(bold=True, color=C_WHITE, size=8)
        else:
            c.font = _font(bold=True, size=8)

    # ── Veri satırları ───────────────────────────────────────────────────────
    for ri, (_, intern) in enumerate(stajyerler_df.iterrows(), 3):
        ws.row_dimensions[ri].height = 18
        static_vals = [
            f"{intern['ad']} {intern['soyad']}",
            intern["staj_gemisi"],
            intern["bolum"],
            intern["calisma_periyodu"],
        ]
        for ci, val in enumerate(static_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = _center()
            c.border    = _thin_border()
            c.font      = _font(size=9)

        izinler_df = get_izinler_for_month(intern["id"], year, month)
        total      = 0

        for d in range(1, num_days + 1):
            ci       = day_col_start + d - 1
            day_date = date(year, month, d)
            status   = get_day_status(day_date, intern["calisma_periyodu"], izinler_df, tr_hols)
            c        = ws.cell(row=ri, column=ci)
            c.alignment = _center()
            c.border    = _thin_border()

            match status:
                case "HAFTA SONU":
                    c.fill = _fill(C_WEEKEND)
                    c.font = _font(color="999999", size=8)
                case "TATİL":
                    c.value = "TATİL"
                    c.fill  = _fill(C_HOLIDAY)
                    c.font  = _font(bold=True, color=C_WHITE, size=7)
                case "-":
                    c.fill = _fill(C_NODAY)
                case "1":
                    c.value = "1"
                    c.fill  = _fill(C_WORKDAY)
                    c.font  = _font(bold=True, color=C_WHITE, size=10)
                    total  += 1
                case "RAPORLU":
                    c.value = "RPL"
                    c.fill  = _fill(C_RAPORLU)
                    c.font  = _font(bold=True, size=8)
                case "RAPORSUZ DEVAMSIZLIK":
                    c.value = "RPSSZ"
                    c.fill  = _fill(C_RAPORSUZ)
                    c.font  = _font(bold=True, size=8)

        tc = ws.cell(row=ri, column=total_col, value=total)
        tc.fill      = _fill(C_TOTAL)
        tc.font      = _font(bold=True, size=10)
        tc.alignment = _center()
        tc.border    = _thin_border()

    # ── Sütun genişlikleri ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    for d in range(1, num_days + 1):
        ws.column_dimensions[get_column_letter(day_col_start + d - 1)].width = 5.5
    ws.column_dimensions[get_column_letter(total_col)].width = 9

    # ── Tatil açıklama satırı ─────────────────────────────────────────────────
    last_data_row = 2 + len(stajyerler_df) + 2
    ws.row_dimensions[last_data_row].height = 14
    month_hols    = {d: n for d, n in tr_hols.items() if d.year == year and d.month == month}
    legend_text   = "Resmi Tatiller: " + ", ".join(
        f"{d.strftime('%d')} {n}" for d, n in sorted(month_hols.items())
    ) if month_hols else "Bu ayda resmi tatil yok."
    ws.cell(row=last_data_row, column=1, value=legend_text).font = _font(size=8, color="595959")

    # Lejant
    lejant_row = last_data_row + 1
    ws.row_dimensions[lejant_row].height = 14
    legends = [
        ("1 → Çalışma günü", C_WORKDAY, C_WHITE),
        ("RPL → Raporlu", C_RAPORLU, "000000"),
        ("RPSSZ → Raporsuz devamsızlık", C_RAPORSUZ, "000000"),
        ("TATİL → Resmi tatil", C_HOLIDAY, C_WHITE),
    ]
    for ci, (txt, bg, fg) in enumerate(legends, 1):
        c = ws.cell(row=lejant_row, column=ci, value=txt)
        c.fill = _fill(bg)
        c.font = _font(color=fg, size=8)

    # ── Dondur ────────────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=3, column=day_col_start)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ══════════════════════════════════════════════════════
#  SAYFA: STAJYER KAYIT
# ══════════════════════════════════════════════════════

def page_kayit():
    st.title("📋 Stajyer Kayıt")
    st.caption("Yeni stajyer ekleyin veya mevcut kayıtları yönetin.")

    # ─── Kayıt formu ──────────────────────────────────────────────────────────
    with st.expander("➕ Yeni Stajyer Ekle", expanded=True):
        with st.form("stajyer_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                ad              = st.text_input("Ad *")
                soyad           = st.text_input("Soyad *")
                okul            = st.text_input("Okul")
                telefon         = st.text_input("Telefon")
            with c2:
                sicil_no        = st.text_input("Sicil No *")
                staj_gemisi     = st.text_input("Staj Gemisi *")
                bolum           = st.selectbox("Bölüm *", BOLUM_OPTIONS)
                calisma_periyodu = st.selectbox("Çalışma Periyodu *", PERIYOT_OPTIONS)

            kaydet = st.form_submit_button("💾 Kaydet", use_container_width=True, type="primary")

        if kaydet:
            if not all([ad.strip(), soyad.strip(), sicil_no.strip(), staj_gemisi.strip()]):
                st.warning("⚠️ Yıldızlı alanlar zorunludur.")
            else:
                try:
                    add_stajyer({
                        "ad": ad.strip(), "soyad": soyad.strip(),
                        "okul": okul.strip(), "telefon": telefon.strip(),
                        "sicil_no": sicil_no.strip(), "staj_gemisi": staj_gemisi.strip(),
                        "bolum": bolum, "calisma_periyodu": calisma_periyodu,
                    })
                    st.success(f"✅ **{ad} {soyad}** başarıyla kaydedildi!")
                    st.rerun()
                except IntegrityError:
                    st.error("❌ Bu sicil numarası zaten kayıtlı.")
                except SQLAlchemyError as exc:
                    st.error(f"❌ Veritabanı hatası: {exc}")

    # ─── Kayıt tablosu ────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📌 Kayıtlı Stajyerler")

    try:
        df = get_all_stajyerler()
    except SQLAlchemyError as exc:
        st.error(f"Veriler yüklenemedi: {exc}")
        return

    if df.empty:
        st.info("Henüz hiç stajyer kaydedilmemiş.")
        return

    # Arama
    arama = st.text_input("🔍 Ara (ad, soyad, gemi, sicil)", placeholder="örn. Ahmet ya da Fırtına")
    if arama:
        mask = (
            df["ad"].str.contains(arama, case=False, na=False) |
            df["soyad"].str.contains(arama, case=False, na=False) |
            df["staj_gemisi"].str.contains(arama, case=False, na=False) |
            df["sicil_no"].str.contains(arama, case=False, na=False)
        )
        df = df[mask]

    col_labels = {
        "id": "ID", "ad": "Ad", "soyad": "Soyad", "okul": "Okul",
        "telefon": "Telefon", "sicil_no": "Sicil No",
        "staj_gemisi": "Gemi", "bolum": "Bölüm", "calisma_periyodu": "Periyot",
    }
    st.dataframe(
        df.rename(columns=col_labels),
        use_container_width=True,
        hide_index=True,
    )

    # Silme
    with st.expander("🗑️ Stajyer Sil"):
        sid_opts = {f"{r['ad']} {r['soyad']} (#{r['id']})": r["id"] for _, r in df.iterrows()}
        if sid_opts:
            sec = st.selectbox("Silmek istenen stajyer", list(sid_opts.keys()))
            if st.button("🗑️ Sil", type="secondary"):
                try:
                    delete_stajyer(sid_opts[sec])
                    st.success("Stajyer ve ilgili izin kayıtları silindi.")
                    st.rerun()
                except SQLAlchemyError as exc:
                    st.error(f"Silme hatası: {exc}")


# ══════════════════════════════════════════════════════
#  SAYFA: GEMİ DASHBOARD
# ══════════════════════════════════════════════════════

def page_dashboard():
    st.title("🚢 Gemi Dashboard")
    st.caption("Gemilere göre stajyer dağılımını inceleyin, filtreleyin.")

    try:
        df = get_all_stajyerler()
    except SQLAlchemyError as exc:
        st.error(f"Veri yüklenemedi: {exc}")
        return

    if df.empty:
        st.info("Henüz kayıtlı stajyer yok. Önce kayıt sayfasından ekleme yapın.")
        return

    # ─── Filtreler ────────────────────────────────────────────────────────────
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        gemiler = ["Tümü"] + sorted(df["staj_gemisi"].dropna().unique().tolist())
        sel_gemi = st.selectbox("Gemi", gemiler)
    with fc2:
        sel_bolum = st.selectbox("Bölüm", ["Tümü"] + BOLUM_OPTIONS)
    with fc3:
        sel_periyot = st.selectbox("Periyot", ["Tümü"] + PERIYOT_OPTIONS)

    filtered = df.copy()
    if sel_gemi    != "Tümü": filtered = filtered[filtered["staj_gemisi"]      == sel_gemi]
    if sel_bolum   != "Tümü": filtered = filtered[filtered["bolum"]             == sel_bolum]
    if sel_periyot != "Tümü": filtered = filtered[filtered["calisma_periyodu"] == sel_periyot]

    # ─── Özet kartlar ─────────────────────────────────────────────────────────
    st.divider()
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("🧑 Toplam Stajyer",  len(filtered))
    m2.metric("⚙️ Makineci",        len(filtered[filtered["bolum"] == "Makineci"]))
    m3.metric("⚓ Güverte",          len(filtered[filtered["bolum"] == "Güverte"]))
    m4.metric("🚢 Aktif Gemi Sayısı", filtered["staj_gemisi"].nunique())

    if filtered.empty:
        st.warning("Seçilen filtrelere uygun stajyer bulunamadı.")
        return

    # ─── Gemi bazlı kart görünümü ─────────────────────────────────────────────
    st.divider()
    st.subheader("🛳️ Gemilere Göre Dağılım")

    ship_stats = (
        filtered
        .groupby("staj_gemisi")
        .apply(lambda x: pd.Series({
            "Toplam":    len(x),
            "Makineci":  (x["bolum"] == "Makineci").sum(),
            "Güverte":   (x["bolum"] == "Güverte").sum(),
        }), include_groups=False)
        .reset_index()
        .rename(columns={"staj_gemisi": "Gemi"})
        .sort_values("Toplam", ascending=False)
    )

    for _, row in ship_stats.iterrows():
        with st.expander(
            f"🚢  **{row['Gemi']}**  —  Toplam: {int(row['Toplam'])} stajyer",
            expanded=True,
        ):
            sc1, sc2, sc3 = st.columns(3)
            sc1.metric("Toplam",   int(row["Toplam"]))
            sc2.metric("Makineci", int(row["Makineci"]))
            sc3.metric("Güverte",  int(row["Güverte"]))

            ship_df = filtered[filtered["staj_gemisi"] == row["Gemi"]][
                ["ad", "soyad", "bolum", "calisma_periyodu", "okul", "sicil_no"]
            ].copy()
            ship_df.columns = ["Ad", "Soyad", "Bölüm", "Periyot", "Okul", "Sicil No"]
            st.dataframe(ship_df, use_container_width=True, hide_index=True)

    # ─── Pasta grafikler ──────────────────────────────────────────────────────
    st.divider()
    st.subheader("📊 Genel Dağılım Grafikleri")

    gc1, gc2 = st.columns(2)
    with gc1:
        st.caption("**Bölüm Dağılımı**")
        bolum_counts = filtered["bolum"].value_counts().reset_index()
        bolum_counts.columns = ["Bölüm", "Sayı"]
        st.bar_chart(bolum_counts.set_index("Bölüm"))
    with gc2:
        st.caption("**Periyot Dağılımı**")
        periyot_counts = filtered["calisma_periyodu"].value_counts().reset_index()
        periyot_counts.columns = ["Periyot", "Sayı"]
        st.bar_chart(periyot_counts.set_index("Periyot"))


# ══════════════════════════════════════════════════════
#  SAYFA: İZİN GİRİŞİ
# ══════════════════════════════════════════════════════

def page_izin():
    st.title("📅 İzin Girişi")
    st.caption("Stajyer izinlerini kaydedin. Sistem otomatik çalışma günü kontrolü yapar.")

    try:
        df = get_all_stajyerler()
    except SQLAlchemyError as exc:
        st.error(f"Veri yüklenemedi: {exc}")
        return

    if df.empty:
        st.warning("Önce stajyer kaydı yapınız.")
        return

    # ─── İzin formu ───────────────────────────────────────────────────────────
    intern_opts = {
        f"{r['ad']} {r['soyad']}  |  {r['staj_gemisi']}  |  #{r['id']}": r["id"]
        for _, r in df.iterrows()
    }

    with st.form("izin_form", clear_on_submit=True):
        sec_intern = st.selectbox("Stajyer Seçin *", list(intern_opts.keys()))

        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            bas_tarih = st.date_input("Başlangıç Tarihi *", value=date.today())
        with fc2:
            bit_tarih = st.date_input("Bitiş Tarihi *",    value=date.today())
        with fc3:
            izin_turu = st.selectbox("İzin Türü *", IZIN_TURU_OPTIONS)

        kaydet = st.form_submit_button("💾 İzni Kaydet", use_container_width=True, type="primary")

    if kaydet:
        if bit_tarih < bas_tarih:
            st.error("❌ Bitiş tarihi, başlangıç tarihinden önce olamaz!")
        else:
            intern_id  = intern_opts[sec_intern]
            intern_row = df[df["id"] == intern_id].iloc[0]
            periyot    = intern_row["calisma_periyodu"]

            has_work, non_work = validate_leave_dates(bas_tarih, bit_tarih, periyot)

            if not has_work:
                st.warning(
                    f"⚠️ **Uyarı:** Seçilen tarih aralığında ({bas_tarih.strftime('%d.%m.%Y')} — "
                    f"{bit_tarih.strftime('%d.%m.%Y')}) bu stajyerin **çalışma günü bulunmamaktadır**.\n\n"
                    f"Stajyerin periyodu: **{periyot}**\n\n"
                    "Bu gün(ler) zaten stajyerin çalışma günü değil. İzin kaydı yapılmadı."
                )
            else:
                if non_work:
                    top5 = ", ".join(non_work[:5])
                    st.info(
                        f"ℹ️ Aşağıdaki günler stajyerin çalışma periyoduna ({periyot}) dahil değil "
                        f"ve izne sayılmayacak:\n**{top5}**"
                        + ("…" if len(non_work) > 5 else "")
                    )
                try:
                    add_izin({
                        "stajyer_id":        intern_id,
                        "baslangic_tarihi":  str(bas_tarih),
                        "bitis_tarihi":      str(bit_tarih),
                        "izin_turu":         izin_turu,
                    })
                    st.success(
                        f"✅ İzin kaydedildi! "
                        f"**{intern_row['ad']} {intern_row['soyad']}** — "
                        f"{bas_tarih.strftime('%d.%m.%Y')} / {bit_tarih.strftime('%d.%m.%Y')} / {izin_turu}"
                    )
                    st.rerun()
                except SQLAlchemyError as exc:
                    st.error(f"❌ Kayıt hatası: {exc}")

    # ─── İzin listesi ─────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📌 Kayıtlı İzinler")

    try:
        iz_df = get_all_izinler()
    except SQLAlchemyError as exc:
        st.error(f"İzinler yüklenemedi: {exc}")
        return

    if iz_df.empty:
        st.info("Henüz kayıtlı izin yok.")
        return

    # Filtre
    fil_gemi = st.selectbox(
        "Gemi filtresi",
        ["Tümü"] + sorted(iz_df["staj_gemisi"].dropna().unique().tolist()),
    )
    show_df = iz_df if fil_gemi == "Tümü" else iz_df[iz_df["staj_gemisi"] == fil_gemi]

    col_labels = {
        "id": "ID", "ad": "Ad", "soyad": "Soyad", "staj_gemisi": "Gemi",
        "baslangic_tarihi": "Başlangıç", "bitis_tarihi": "Bitiş", "izin_turu": "Tür",
    }
    st.dataframe(show_df.rename(columns=col_labels), use_container_width=True, hide_index=True)

    # Silme
    with st.expander("🗑️ İzin Sil"):
        if not iz_df.empty:
            izin_opts = {
                f"#{r['id']} | {r['ad']} {r['soyad']} | {r['baslangic_tarihi']} → {r['bitis_tarihi']}": r["id"]
                for _, r in iz_df.iterrows()
            }
            sec_izin = st.selectbox("Silmek istenen izin", list(izin_opts.keys()))
            if st.button("🗑️ İzni Sil", type="secondary"):
                try:
                    delete_izin(izin_opts[sec_izin])
                    st.success("İzin kaydı silindi.")
                    st.rerun()
                except SQLAlchemyError as exc:
                    st.error(f"Silme hatası: {exc}")


# ══════════════════════════════════════════════════════
#  SAYFA: PUANTAJ ÇIKTISI
# ══════════════════════════════════════════════════════

def page_puantaj():
    st.title("📊 Puantaj Çıktısı")
    st.caption("Aylık çalışma tablosunu görüntüleyin ve .xlsx olarak indirin.")

    # ─── Filtre çubuğu ────────────────────────────────────────────────────────
    pc1, pc2, pc3 = st.columns([1, 1, 2])
    yil_secenekleri = list(range(2022, date.today().year + 3))
    with pc1:
        sec_yil = st.selectbox("Yıl", yil_secenekleri, index=yil_secenekleri.index(date.today().year))
    with pc2:
        sec_ay_adi = st.selectbox("Ay", AY_ADLARI, index=date.today().month - 1)
        sec_ay     = AY_ADLARI.index(sec_ay_adi) + 1

    try:
        df = get_all_stajyerler()
    except SQLAlchemyError as exc:
        st.error(f"Veri yüklenemedi: {exc}")
        return

    with pc3:
        gemiler    = ["Tümü"] + sorted(df["staj_gemisi"].dropna().unique().tolist()) if not df.empty else ["Tümü"]
        sel_gemi_p = st.selectbox("Gemi Filtresi", gemiler)

    if df.empty:
        st.warning("Önce stajyer kaydı yapınız.")
        return

    if sel_gemi_p != "Tümü":
        df = df[df["staj_gemisi"] == sel_gemi_p]

    if df.empty:
        st.info("Seçilen gemi için stajyer bulunamadı.")
        return

    tr_hols  = get_tr_holidays(sec_yil)
    num_days = calendar.monthrange(sec_yil, sec_ay)[1]

    # ─── Tatil bilgisi ────────────────────────────────────────────────────────
    month_hols = {d: n for d, n in tr_hols.items() if d.year == sec_yil and d.month == sec_ay}
    if month_hols:
        hol_txt = "  ·  ".join(f"**{d.strftime('%d')}** {n}" for d, n in sorted(month_hols.items()))
        st.info(f"🎌 Resmi tatiller: {hol_txt}")
    else:
        st.caption("Bu ayda resmi tatil yok.")

    # ─── Ekran tablosu ────────────────────────────────────────────────────────
    st.subheader(f"📅 {sec_ay_adi} {sec_yil} — Puantaj")

    rows = []
    for _, intern in df.iterrows():
        izinler_df = get_izinler_for_month(intern["id"], sec_yil, sec_ay)
        row = {
            "Ad Soyad": f"{intern['ad']} {intern['soyad']}",
            "Gemi":     intern["staj_gemisi"],
            "Bölüm":    intern["bolum"],
        }
        total = 0
        for d in range(1, num_days + 1):
            day_date = date(sec_yil, sec_ay, d)
            status   = get_day_status(day_date, intern["calisma_periyodu"], izinler_df, tr_hols)
            col_key  = f"{d}"
            match status:
                case "HAFTA SONU":           row[col_key] = "〇"
                case "TATİL":                row[col_key] = "🎌"
                case "-":                    row[col_key] = ""
                case "1":                    row[col_key] = "✅"; total += 1
                case "RAPORLU":              row[col_key] = "📋"
                case "RAPORSUZ DEVAMSIZLIK": row[col_key] = "❌"
                case _:                      row[col_key] = ""
        row["Toplam"] = total
        rows.append(row)

    result_df = pd.DataFrame(rows)
    st.dataframe(result_df, use_container_width=True, hide_index=True)
    st.caption(
        "**Lejant:** ✅ Çalışma günü  ·  🎌 Resmi tatil  ·  〇 Hafta sonu  ·  "
        "📋 Raporlu izin  ·  ❌ Raporsuz devamsızlık  ·  Boş → Periyot dışı gün"
    )

    # ─── Excel indirme ────────────────────────────────────────────────────────
    st.divider()
    st.subheader("⬇️ Excel İndir")

    with st.spinner("Excel hazırlanıyor…"):
        try:
            excel_buf = create_excel(sec_yil, sec_ay, df)
            st.download_button(
                label=f"📥  {sec_ay_adi} {sec_yil} Puantaj (.xlsx)",
                data=excel_buf,
                file_name=f"Puantaj_{sec_yil}_{sec_ay:02d}_{sel_gemi_p.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        except Exception as exc:
            st.error(f"Excel oluşturulamadı: {exc}")


# ══════════════════════════════════════════════════════
#  SIDEBAR & YÖNLENDIRME
# ══════════════════════════════════════════════════════

PAGES = {
    "📋 Stajyer Kayıt":  page_kayit,
    "🚢 Gemi Dashboard": page_dashboard,
    "📅 İzin Girişi":    page_izin,
    "📊 Puantaj Çıktısı": page_puantaj,
}


def main():
    # CSS ince dokunuş
    st.markdown(
        """
        <style>
        [data-testid="stSidebarNav"] { display: none; }
        .stMetric { background: #f0f4f8; border-radius: 8px; padding: 8px; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    init_db()

    st.sidebar.image(
        "https://upload.wikimedia.org/wikipedia/commons/thumb/3/3d/Anchor_pictogram.svg/240px-Anchor_pictogram.svg.png",
        width=60,
    )
    st.sidebar.title("Stajyer Takip\nSistemi")
    st.sidebar.divider()

    sel = st.sidebar.radio("Sayfa Seçin", list(PAGES.keys()), label_visibility="collapsed")
    st.sidebar.divider()

    # Mini istatistik
    try:
        df_all = get_all_stajyerler()
        st.sidebar.metric("Toplam Stajyer", len(df_all))
        if not df_all.empty:
            st.sidebar.metric("Aktif Gemi",   df_all["staj_gemisi"].nunique())
    except Exception:
        pass

    st.sidebar.divider()
    st.sidebar.caption("🚢 v1.0 | Stajyer Takip")

    PAGES[sel]()


if __name__ == "__main__":
    main()
